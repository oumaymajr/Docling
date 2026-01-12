import csv
import re
import logging
from pathlib import Path
from openpyxl import load_workbook, Workbook
from docling.document_converter import DocumentConverter
from docling_core.types.doc import ImageRefMode

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s"
)

log = logging.getLogger("TabularDoclingPipeline")

class TabularDoclingPipeline:
    TABLE_REGEX = re.compile(
        r"(\|.+\|\n\|[-:\s|]+\|\n(?:\|.*\|\n?)+)",
        re.MULTILINE
    )

    def __init__(self):
        self.converter = DocumentConverter()


    # ======================================================
    #                     XLSX
    # ======================================================

    def convert_xlsx(self, xlsx_path: Path) -> str:
        log.info(f"Conversion XLSX : {xlsx_path.name}")
        tmp_xlsx = xlsx_path.with_suffix(".values.xlsx")
        log.info("Conversion XLSX → valeurs uniquement")
        wb_src = load_workbook(xlsx_path, data_only=True)
        wb_dst = wb_src
        for ws in wb_dst.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f":  # formule
                        log.debug(f"Formule remplacée par valeur : {cell.coordinate}")

        wb_dst.save(tmp_xlsx)
        result = self.converter.convert(str(tmp_xlsx))
        output_md = xlsx_path.with_suffix(".md")
        result.document.save_as_markdown(
            output_md,
            image_mode=ImageRefMode.EMBEDDED
        )

        log.info(" Markdown Docling généré")
        md_text = output_md.read_text(encoding="utf-8")
        wb = load_workbook(xlsx_path, read_only=True)
        sheet_names = wb.sheetnames
        tables = self.TABLE_REGEX.findall(md_text)
        rebuilt = []
        table_idx = 0
        for sheet in sheet_names:
            if table_idx >= len(tables):
                break
            rebuilt.append(f"## {sheet}")
            rebuilt.append(tables[table_idx].strip())
            table_idx += 1
        md_text = "\n\n".join(rebuilt)
        tmp_xlsx.unlink(missing_ok=True)
        output_md.unlink(missing_ok=True)

        return md_text

    # ======================================================
    #                    CSV
    # ======================================================

    def detect_separator(self, path: Path) -> str:
        with path.open(encoding="utf-8-sig", errors="ignore") as f:
            line = f.readline()
        return "," if line.count(",") > line.count(";") else ";"

    def clean_cell(self, value: str) -> str:
        if not value:
            return ""
        value = re.sub(r"[|\n\r\t]", " ", value)
        value = re.sub(r"\s+", " ", value)
        return value.strip()

    def convert_csv(self, csv_path: Path) -> str:
        sep = self.detect_separator(csv_path)
        with csv_path.open(encoding="utf-8-sig", errors="ignore") as f:
            reader = list(csv.reader(f, delimiter=sep))

        if not reader:
            raise RuntimeError("CSV vide")

        header = [self.clean_cell(h) for h in reader[0]]
        body = reader[1:]

        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * len(header)) + " |"
        ]

        for row in body:
            row = row[:len(header)]
            row += [""] * (len(header) - len(row))
            row = [self.clean_cell(c) for c in row]
            lines.append("| " + " | ".join(row) + " |")

        return "\n".join(lines)



